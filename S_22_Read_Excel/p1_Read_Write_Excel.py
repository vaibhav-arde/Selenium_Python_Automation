import openpyxl

# Load the workbook and select the active sheet
file_path = "S_22_Read_Excel/data.xlsx"
workbook = openpyxl.load_workbook(file_path)
sheet = workbook.active
cell = sheet.cell(row=1, column=2)
print(cell.value)
sheet.cell(row= 5, column= 1).value = "WriteData"
print(sheet.cell(row= 5, column= 1).value)

print("Max Rows",sheet.max_row)
print("Max Columns",sheet.max_column)

print(sheet['A5'].value)
# # Print all values in column A (e.g., Username column)
print("Print values in Column A")
for cell in sheet['A']:  # Replace 'A' with the desired column letter
    print(cell.value)

print("Print values in Row 1")
for cell in sheet[1]:  # Replace 'A' with the desired column letter
    print(cell.value)
    
print("Print values in Column A : ")
for i in range(2, sheet.max_row+1):  # Replace 'A' with the desired column letter
    print(sheet.cell(row=i, column=1).value)
    
print("Print values in Row : ")
for i in range(1, sheet.max_column+1):  # Replace 'A' with the desired column letter
    print(sheet.cell(row=1, column=i).value)
    
print("All data :")
for i in range(2, sheet.max_row + 1):
    for j in range(1, sheet.max_column +1):
        print(sheet.cell(i,j).value)
        
print("All row for selected username :")
for i in range(2, sheet.max_row + 1):
    if sheet.cell(i,1).value == "user2":
        for j in range(1, sheet.max_column +1):
            print(sheet.cell(i,j).value)